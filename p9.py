#######################################################################################
# PH - Performance History
# TH - Trade History
# DB - Dashboard
#######################################################################################

import csv
import os
import openpyxl
from p2 import eq_opt_ltp
from p8 import get_internet_time
import time

def read_csv(file_path):
    data = []
    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            data.append(row)
    return data
    #IO0955,,SBIN,860,CE,26-Jun-2025,S,6000,6.05,,

def update_dshboard():
    trades=read_csv("Input/IO0955.csv")

    wb=openpyxl.load_workbook('Output/testing.xlsx')
    sheet=wb['Sheet1']

    th=openpyxl.load_workbook('Output/TH.xlsx')

    cr=1
    skip=[]

    for row in trades:
        # print(row[2],int(row[3]),row[5],row[4])

        #eq, stk, exp, typ >>> [eq, uv, stk, exp, typ, ltp, iv, oi, coi, ttv]
        # print(f">>>> {eq_opt_ltp(row[2],int(row[3]),row[5],row[4])}")

        data=eq_opt_ltp(row[2],int(row[3]),row[5],row[4])

        cr+=1

        #################PH Loading#################

        #Acc
        sheet.cell(cr,1).value=row[0]
        #Date
        sheet.cell(cr,2).value=row[1]
        #Trade
        trade=row[2]+" "+row[5]+" "+row[3]+" "+row[4]
        sheet.cell(cr,3).value=trade
        #UV
        sheet.cell(cr,4).value=float(data[1])
        #Buy/Sell
        sheet.cell(cr,5).value="Sell" if row[6] == "S" else "Buy"
        #Avg.
        avg=round(float(row[8]),2)
        sheet.cell(cr,6).value=avg
        #Qty
        qty=int(row[7])
        sheet.cell(cr,7).value=qty
        #LTP
        ltp=round(float(data[5]),2)
        sheet.cell(cr,8).value=ltp
        #Close Date
        sheet.cell(cr,9).value=row[9]
        #Close Price
        sheet.cell(cr,10).value=round(float(row[10]),2) if row[10] != '' else row[10]
        #P/L
        sheet.cell(cr,12).value=round(float(qty*(avg-ltp)*(1 if row[6] == "S" else -1)),2)

        #################TH Loading#################

        if trade in th.sheetnames:
            if trade not in skip:
                # print(trade,skip)
                skip.append(trade)
                thsheet = th[trade]
                thsheet.append([get_internet_time()]+data)
            else:
                pass
        else:
            th['Sheet1'].append([trade])
            thsheet = th.create_sheet(trade)
            thsheet.append(['timestamp', 'eq', 'uv', 'stk', 'exp', 'typ', 'ltp', 'iv', 'oi', 'coi', 'ttv', '=HYPERLINK("#Sheet1!A1", "Go to Sheet: Home")'])
            thsheet.append([get_internet_time()]+data)


    wb.save('Output/testing.xlsx')
    wb.close()

    th.save('Output/TH.xlsx')
    th.close()

# update_dshboard()

def take_snap():
    wb = openpyxl.load_workbook('Output/testing.xlsx', data_only=True)
    sheet = wb['Sheet1']


    # tr = [row[0].value for row in sheet["J2:J9"]]
    # pl = [row[0].value for row in sheet["J2:J9"]]

    # Step 1: Read column J from Sheet1 starting at J2 until empty
    tr = []
    pl = []
    row = 2
    while True:
        v1 = sheet.cell(row=row, column=1).value+" "+sheet.cell(row=row, column=3).value if  sheet.cell(row=row, column=1).value is not None else sheet.cell(row=row, column=1).value # C = column 3
        v2 = sheet.cell(row=row, column=12).value  if sheet.cell(row=row, column=9).value is None else 'Closed' # L = column 12
        if v1 is None:
            break
        # print(v1,v2)
        tr.append(v1)
        pl.append(v2)
        row += 1

    wb.close()

    return tr,pl

# print(take_snap())

def dashboard_snap(header, data):
    # print(header,data)
    file_exists = os.path.exists('Output/PH.xlsx')

    # Step 1: Load or create workbook
    if file_exists:
        wb = openpyxl.load_workbook('Output/PH.xlsx')
        sheet = wb['Sheet1']
        if sheet.max_column != len(header)+2:
            [sheet.cell(row=1, column=i+1, value=val) for i, val in enumerate(['Timestamp','Total']+header)]
            print("header changed")

    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Sheet1'

        # Step 2: Write header if new file
        sheet.append(['Timestamp','Total']+header)

    # Step 3: Append data
    sheet.append([get_internet_time()]+[sum(x for x in data if isinstance(x, (int, float)))]+data)
    print(sum(x for x in data if isinstance(x, (int, float))))

    # Step 4: Save
    wb.save('Output/PH.xlsx')
    wb.close()
    

# print(dashboard_snap(*take_snap()))

def call():

    while True:
        print(get_internet_time())
        update_dshboard() 
        dashboard_snap(*take_snap())
        print("Success")
        time.sleep(300)  # Sleep for 5 minutes



# Start periodic call
#call()